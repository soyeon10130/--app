import streamlit as st
import pandas as pd
import io
import datetime
import re
from datetime import timedelta, time
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

KST_OFFSET = timedelta(hours=9)
SET1_DEST  = {'DAD','BKK','HKG','NRT'}
SET2_DEST  = {'DAC','LAX','EWR','SFO','IAD'}
SET3P_DEST = {'HNL'}
INSTR_DC   = {'LIP','LCP','DLCP','I','I*'}
INSTR_EXCL = {'강용학','김문배','박충근','박형득','서세규'}

# ═══════════════════════════════════════════
# 램프리턴/불이착 자동 보정
# ═══════════════════════════════════════════
RAMP_RETURN_GAP_WINDOW_H = 6.0   # 회항 도착~재출발 ATD 간격이 이보다 크면 별도 비행(다음날 등)으로 보고 병합 안함

def normalize_ramp_returns(df):
    """
    FltReport 원본에서 램프리턴/불이착(From==To 회항편)을 자동 감지해
    운항정산 계산 전에 보정한다.

    처리 규칙 (연장시간 = ATD~ATA 전체 경과시간 기준, 같은 편명 램프리턴에만 적용):
      1) From==To 행(회항편)을 찾는다.
      2) 같은 편명 · 같은 출발지에서 그 이후에 실제로 목적지가 다른(To!=From)
         재출발 행을 찾는다 (회항 도착 후 RAMP_RETURN_GAP_WINDOW_H 시간 이내).
      3) 회항 시간 길이와 상관없이(이륙 전 회항이든 실제 공중 회항이든) 두 행을 하나로 병합:
         ATD는 회항편의 최초 출발 시각, ATA·Bl Hrs·운항 C/I·운항 C/O·To는 재출발편 값을 사용.
         → 연장시간(ATD~ATA 경과시간) 계산 시 회항에 걸린 시간까지 전부 포함됨.
      4) 재출발편을 찾지 못하거나 간격이 너무 크면(다음날 새 비행 등) 원본을 그대로 둔다
         (예: SFO→SFO 램프리턴 후 다음날 새 편으로 출발한 케이스는 별도 duty로 보고 병합하지 않음)

    ※ 편명이 다른 당일 왕복(예: NRT 731 나가고 732 들어오는 경우)은 이 병합 대상이 아니며,
    기존 방식대로 각 편명별로 개별 계산된다.

    ※ 주의: 목적지가 아예 다른 공항으로 바뀌는 '진짜 다이버트'(예: 예정 도착지가 ICN인데
    실제로는 다른 공항에 착륙한 뒤 별도의 이동편이 필요한 경우)는 원본 FltReport에
    해당 이동편 데이터 자체가 없는 경우가 있어 자동 보정 대상에서 제외한다.
    이런 경우는 여전히 수동 확인/보정이 필요하다.
    """
    d = df.copy().reset_index(drop=True)

    date_dt = pd.to_datetime(d['Date'])

    def to_dt(i, t):
        if pd.isna(t):
            return None
        base = date_dt.iloc[i]
        if isinstance(t, datetime.time):
            return datetime.datetime.combine(base.date(), t)
        return None

    d['_atd_dt'] = [to_dt(i, d.at[i, 'ATD']) for i in d.index]
    d['_ata_dt'] = [to_dt(i, d.at[i, 'ATA']) for i in d.index]
    for i in d.index:
        a, b = d.at[i, '_atd_dt'], d.at[i, '_ata_dt']
        if a is not None and b is not None and b < a:
            d.at[i, '_ata_dt'] = b + timedelta(days=1)

    d['_flight_s'] = d['Flight'].astype(str)
    d = d.sort_values('_atd_dt', kind='stable').reset_index(drop=True)

    drop_idx = set()
    replace_rows = {}
    merge_log = []   # (날짜, 편명, 출발지, 설명) 기록용

    return_rows = d[d['From'] == d['To']].index.tolist()

    for i in return_rows:
        if i in drop_idx:
            continue
        row = d.loc[i]
        flight = row['_flight_s']
        origin = row['From']
        bl_hrs = row['Bl Hrs']
        bl_h = (bl_hrs.hour + bl_hrs.minute / 60 + bl_hrs.second / 3600) if isinstance(bl_hrs, datetime.time) else 0.0
        if row['_atd_dt'] is None:
            continue

        candidates = d[
            (d['_flight_s'] == flight) &
            (d['From'] == origin) &
            (d['To'] != origin) &
            (d['_atd_dt'].notna()) &
            (d['_atd_dt'] > row['_atd_dt'])
        ].sort_values('_atd_dt')

        if candidates.empty:
            continue

        nxt = candidates.iloc[0]
        j = nxt.name
        if row['_ata_dt'] is None or nxt['_atd_dt'] is None:
            continue
        gap_hours = (nxt['_atd_dt'] - row['_ata_dt']).total_seconds() / 3600

        if gap_hours > RAMP_RETURN_GAP_WINDOW_H:
            continue  # 다음날 새 비행 등, 병합 대상 아님

        date_label = row['_atd_dt'].strftime("%m/%d")

        # 회항 시간 길이와 무관하게 항상 병합 (ATD는 회항편, 나머지는 재출발편)
        drop_idx.add(i)
        drop_idx.add(j)
        merged = nxt.copy()
        merged['ATD'] = row['ATD']
        if pd.notna(row['STD']):
            merged['STD'] = row['STD']
        replace_rows[j] = merged
        merge_log.append((date_label, flight, origin, f"램프리턴(회항 Bl {bl_h*60:.0f}분) 자동 병합 — 연장시간에 회항 경과시간 포함"))

    out_rows = []
    for idx, r in d.iterrows():
        if idx in replace_rows:
            out_rows.append(replace_rows[idx])
        elif idx in drop_idx:
            continue
        else:
            out_rows.append(r)

    result = pd.DataFrame(out_rows).drop(columns=['_atd_dt', '_ata_dt', '_flight_s']).reset_index(drop=True)
    return result, merge_log

# ═══════════════════════════════════════════
# 공통 유틸
# ═══════════════════════════════════════════
def fmt_hhmm(h):
    if h is None or (isinstance(h, float) and pd.isna(h)) or h == 0:
        return "-"
    neg = h < 0; h = abs(h)
    hours = int(h); mins = round((h - hours) * 60)
    if mins == 60: hours += 1; mins = 0
    return f"{'-' if neg else ''}{hours:02d}:{mins:02d}"

def fmt_time(t):
    if t is None or (isinstance(t, float) and pd.isna(t)): return "-"
    if isinstance(t, (time, datetime.datetime)): return t.strftime("%H:%M")
    return str(t)

def parse_hhmm(val):
    if pd.isna(val): return 0.0
    val = str(val).strip()
    if ":" in val:
        p = val.split(":")
        try: return int(p[0]) + int(p[1]) / 60
        except: return 0.0
    try: return float(val)
    except: return 0.0

def td_to_hours(val):
    if isinstance(val, datetime.timedelta): return val.total_seconds() / 3600
    return 0.0

def make_border(color="BBBBBB"):
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def style_hdr(ws, row, headers, bg="1F4E79", height=20):
    bd = make_border()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hb = PatternFill("solid", fgColor=bg)
    ct = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font=hf; c.fill=hb; c.alignment=ct; c.border=bd
    ws.row_dimensions[row].height = height

# ═══════════════════════════════════════════
# 파일1: DHC / DAYOFF 파싱
# ═══════════════════════════════════════════
def parse_dhc_file(uploaded):
    df = pd.read_excel(uploaded)
    data = df.iloc[4:].copy()
    data.columns = ["Crew Code","Position","Name","Counter","Mar26","Total"]
    data = data[data["Counter"] != "Counter"].copy()
    block  = data[data["Counter"]=="Block"][["Crew Code","Mar26"]].copy()
    dayoff = data[data["Counter"]=="dayoff"][["Crew Code","Mar26"]].copy()
    dhc    = data[data["Counter"]=="DHC"][["Crew Code","Mar26"]].copy()
    block.columns=["Crew Code","Block"]; block["Block"]=block["Block"].apply(parse_hhmm)
    dayoff.columns=["Crew Code","Dayoff"]; dayoff["Dayoff"]=pd.to_numeric(dayoff["Dayoff"],errors="coerce").fillna(0).astype(int)
    dhc.columns=["Crew Code","DHC"]; dhc["DHC"]=dhc["DHC"].apply(parse_hhmm)
    return block.merge(dayoff, on="Crew Code").merge(dhc, on="Crew Code")

# ═══════════════════════════════════════════
# 파일: OBCA.xlsx 전용 파싱 (OBCA/OBFO Block Hours 합산)
# ═══════════════════════════════════════════
def parse_obca_file(uploaded):
    """
    OBCA.xlsx 전용 리포트 파싱 (Crew Code / AC Type / Position / Block Hours ... 형식,
    8번째 행이 헤더). Position이 OBFO 또는 OBCA인 행의 Block Hours(timedelta)를
    승무원별로 합산하여 반환.
    """
    df = pd.read_excel(uploaded)
    data = df.iloc[6:].copy()
    data.columns = ["Crew Code","AC Type","Position","Block Hours","Cruise Time","Sectors","Valid From","Valid To"]
    data = data[data["Crew Code"] != "Crew Code"].copy()
    data["Crew Code"] = data["Crew Code"].ffill()
    ob = data[data["Position"].isin(["OBFO","OBCA"])].copy()
    ob["OB_hrs"] = ob["Block Hours"].apply(td_to_hours)
    ob_sum = ob.groupby("Crew Code")["OB_hrs"].sum().reset_index()
    ob_sum.columns=["Crew Code","OBCA_OBFO"]
    return ob_sum

def merge_ob_sums(*ob_dfs):
    """여러 출처(전체 승무원 Roster 자동감지 + 별도 OBCA.xlsx)의 OBCA/OBFO 합산분을
    승무원(Crew Code)별로 합쳐서 하나의 DataFrame으로 반환. 이중 계산 방지를 위해
    각 출처는 서로 겹치지 않는 것을 전제로 단순 합산한다."""
    frames = [d for d in ob_dfs if d is not None and not d.empty]
    if not frames:
        return pd.DataFrame(columns=["Crew Code","OBCA_OBFO"])
    combined = pd.concat(frames, ignore_index=True)
    return combined.groupby("Crew Code")["OBCA_OBFO"].sum().reset_index()

def calc_summary(base_df, ob_df=None):
    df = base_df.copy()
    if ob_df is not None and not ob_df.empty:
        df = df.merge(ob_df, on="Crew Code", how="left")
        df["OBCA_OBFO"] = df["OBCA_OBFO"].fillna(0)
    else:
        df["OBCA_OBFO"] = 0.0
    df["DHC_50"]    = df["DHC"] * 0.5
    df["Total_Flt"] = (df["Block"] + df["DHC_50"] - df["OBCA_OBFO"]).clip(lower=0)
    df["Dayoff_Under8"] = df["Dayoff"] < 8
    return df

# ═══════════════════════════════════════════
# Roster 파싱 — 교관수당 + DH/OBCA/OBFO 자동 감지 통합
# ═══════════════════════════════════════════
AIRPORT_UTC = {
    'ICN': 9, 'NRT': 9, 'HKG': 8, 'BKK': 7, 'DAD': 7,
    'LAX': -8, 'EWR': -5, 'SFO': -8, 'IAD': -5, 'DAC': 6, 'HNL': -10,
}

DOMESTIC_REPOSITION = {'ICN', 'GMP'}  # 인천 커퓨(23:00~06:00) 회피용 김포 착륙 후 재배치 등 국내 구간

def classify_route(from_val, to_val):
    for v in [from_val, to_val]:
        v = str(v).strip().upper()
        if v in DOMESTIC_REPOSITION: continue
        if v in SET1_DEST:  return "1set"
        if v in SET2_DEST:  return "2set"
        if v in SET3P_DEST: return "3P"
    return None

def build_flight_set_map(uploaded):
    """
    Roster 전체를 훑어서 '편명 → Set구분(1set/2set/3P)' 매핑을 만든다.
    같은 편명은 항상 같은 노선을 운항하므로, GMP↔ICN 같은 국내 재배치 구간처럼
    From/To만으로 노선을 판단할 수 없는 행에 대한 fallback으로 사용한다.
    (예: 인천 커퓨로 김포 착륙 후 ICN으로 재배치하는 짧은 구간)
    """
    raw, name_indices = _get_crew_sections(uploaded)
    votes = {}  # flight_no(str) -> {set_type: count}
    for idx_i, name_idx in enumerate(name_indices[:-1]):
        data = _parse_crew_data(raw, name_idx, name_indices[idx_i + 1])
        if data is None:
            continue
        flights = data[data["Activity"].apply(is_actual_flight)]
        for _, row in flights.iterrows():
            from_val = str(row["From"]).strip() if not pd.isna(row["From"]) else ""
            to_val   = str(row["To"]).strip()   if not pd.isna(row["To"])   else ""
            set_type = classify_route(from_val, to_val)
            if set_type is None:
                continue
            m = re.search(r'\d+', str(row["Activity"]))
            if not m:
                continue
            flight_no = m.group()
            votes.setdefault(flight_no, {}).setdefault(set_type, 0)
            votes[flight_no][set_type] += 1
    return {flt: max(counts, key=counts.get) for flt, counts in votes.items()}

def local_to_kst(date_str, time_str, from_city):
    try:
        base = datetime.datetime.strptime(str(date_str).strip(), "%d%b%y")
    except:
        return None
    s = str(time_str).strip() if not pd.isna(time_str) else ""
    if not s or s == "nan":
        return datetime.datetime(base.year, base.month, base.day, 0, 0)
    next_day = "+1" in s
    prev_day = "-1" in s
    s_clean = s.replace("+1", "").replace("-1", "").strip()
    try:
        h = int(s_clean[:2]); m = int(s_clean[2:4])
    except:
        return datetime.datetime(base.year, base.month, base.day, 0, 0)
    local_dt = datetime.datetime(base.year, base.month, base.day, h, m)
    if next_day: local_dt += timedelta(days=1)
    if prev_day: local_dt -= timedelta(days=1)
    offset = AIRPORT_UTC.get(str(from_city).strip().upper(), 9)
    return local_dt + timedelta(hours=(9 - offset))

def calc_instr_hrs(blhr_h, set_type):
    if set_type == "1set":  return blhr_h
    if set_type == "2set":  return blhr_h / 2
    if set_type == "3P":    return blhr_h / 3
    return 0.0

def is_actual_flight(activity):
    if pd.isna(activity): return False
    return str(activity).strip().upper().startswith("YP")

def _get_crew_sections(uploaded):
    """Roster 파일에서 승무원별 섹션 인덱스 반환"""
    df = pd.read_excel(uploaded)
    raw = df.copy()
    name_indices = raw[raw.iloc[:,0].astype(str).str.match(r'^[가-힣]{2,5}:$', na=False)].index.tolist()
    name_indices.append(len(raw))
    return raw, name_indices

def _parse_crew_data(raw, name_idx, next_idx):
    """승무원 한 명의 스케줄 블록을 DataFrame으로 파싱
    - 15컬럼: Date/Pairing/DC/Pos/CI_L/CO_L/Activity/From/Start_L/To/Finish_L/AC_Hotel/BH/FDP/Blhr (전체 Roster)
    - 14컬럼: Date/Pairing/DC/CI_L/CO_L/Activity/From/Start_L/To/Finish_L/AC_Hotel/BH/FDP/Blhr (교관수당 Roster)
    """
    hdr_rows = raw.iloc[name_idx:next_idx][raw.iloc[name_idx:next_idx, 0] == "Date"].index
    if len(hdr_rows) == 0:
        return None
    hdr_idx = hdr_rows[0]
    data = raw.iloc[hdr_idx+1:next_idx].copy()
    ncols = len(data.columns)
    if ncols >= 15:
        data.columns = ["Date","Pairing","DC","Pos","CI_L","CO_L","Activity",
                        "From","Start_L","To","Finish_L","AC_Hotel","BH","FDP","Blhr"] + list(data.columns[15:])
    else:
        data.columns = ["Date","Pairing","DC","CI_L","CO_L","Activity",
                        "From","Start_L","To","Finish_L","AC_Hotel","BH","FDP","Blhr"] + list(data.columns[14:])
        data["Pos"] = ""  # Pos 컬럼 없으면 빈값으로 추가
    data = data.reset_index(drop=True)
    data["Pairing_ff"] = data["Pairing"].ffill()
    data["Date_ff"]    = data["Date"].ffill()

    # 사이클 구분: Pairing이 재등장할 때 이전 사이클에 오는편(짝수 편명)이 있으면 새 사이클
    group_id = 0
    last_pairing = None
    last_had_return = False  # 이전 사이클에 짝수(오는편) 편명이 있었는지
    group_ids = []
    pending_flights = set()   # 현재 사이클의 편명 번호들

    for i, row in data.iterrows():
        pairing = row["Pairing"]
        activity = str(row["Activity"]).strip() if not pd.isna(row["Activity"]) else ""
        is_flight = activity.upper().startswith("YP")

        if not pd.isna(pairing):
            if last_pairing != str(pairing):
                # Pairing 값 자체가 바뀜 → 무조건 새 사이클
                group_id += 1
                pending_flights = set()
                last_had_return = False
            else:
                # 같은 Pairing 재등장 → 이전 사이클에 짝수편명 있었으면 새 사이클
                if last_had_return:
                    group_id += 1
                    pending_flights = set()
                    last_had_return = False
            last_pairing = str(pairing)

        if is_flight:
            m = re.search(r'\d+', activity)
            if m:
                num = int(m.group())
                pending_flights.add(num)
                if num % 2 == 0:  # 짝수 = 오는편 완료
                    last_had_return = True

        group_ids.append(group_id)

    data["group_id"] = group_ids
    first_valid = data["Pairing_ff"].first_valid_index()
    if first_valid is None:
        return None
    return data[(data.index >= first_valid) & (data["group_id"] > 0)]

# ── 교관수당 전용 파싱 (기존 Roster — LIP/LCP/DLCP 포지션 추출) ─────────────
def parse_roster_file(uploaded, target_month=None, target_year=None, flight_set_map=None):
    """교관 수당 파싱 전용"""
    raw, name_indices = _get_crew_sections(uploaded)
    detail_rows = []
    flight_set_map = flight_set_map or {}

    for idx_i, name_idx in enumerate(name_indices[:-1]):
        crew_name = str(raw.iloc[name_idx, 0]).replace(":", "").strip()
        if crew_name in INSTR_EXCL:
            continue
        data = _parse_crew_data(raw, name_idx, name_indices[idx_i + 1])
        if data is None:
            continue

        for gid, grp in data.groupby("group_id"):
            if grp["Pairing_ff"].isna().all(): continue
            if not grp["DC"].isin(INSTR_DC).any(): continue

            all_flights = grp[grp["Activity"].apply(is_actual_flight)].copy()
            all_flights["Blhr_h"] = all_flights["Blhr"].apply(parse_hhmm)
            all_flights = all_flights[all_flights["Blhr_h"] > 0]

            # INSTR_DC 붙은 편명 번호의 홀짝 쌍만 포함
            instr_nums = set()
            for _, fr in all_flights[all_flights["DC"].isin(INSTR_DC)].iterrows():
                m = re.search(r'\d+', str(fr["Activity"]))
                if not m: continue
                num = int(m.group())
                instr_nums.add(num)
                instr_nums.add(num + 1 if num % 2 == 1 else num - 1)

            def in_pair(activity):
                m = re.search(r'\d+', str(activity))
                return int(m.group()) in instr_nums if m else False

            flights = all_flights[all_flights["Activity"].apply(in_pair)]

            for _, row in flights.iterrows():
                from_val = str(row["From"]).strip() if not pd.isna(row["From"]) else ""
                to_val   = str(row["To"]).strip()   if not pd.isna(row["To"])   else ""
                set_type = classify_route(from_val, to_val)
                if set_type is None:
                    # From/To가 둘 다 국내(예: 인천 커퓨로 인한 GMP↔ICN 재배치)라 판단 불가한 경우,
                    # 같은 편명의 다른 정상 구간들로부터 만든 매핑으로 보정
                    m = re.search(r'\d+', str(row["Activity"]))
                    set_type = flight_set_map.get(m.group()) if m else None
                if set_type is None: continue

                kst_dt = local_to_kst(row["Date_ff"], row["Start_L"], from_val)
                if kst_dt is None: continue
                if target_month and target_year:
                    if kst_dt.month != target_month or kst_dt.year != target_year:
                        continue

                blhr_h  = row["Blhr_h"]
                instr_h = calc_instr_hrs(blhr_h, set_type)
                dc_val  = str(row["DC"]).strip() if not pd.isna(row["DC"]) else "-"

                detail_rows.append({
                    "이름":       crew_name,
                    "날짜(KST)":  kst_dt.strftime("%m/%d"),
                    "DC":         dc_val,
                    "편명":       str(row["Activity"]).strip(),
                    "From":       from_val,
                    "To":         to_val,
                    "Set구분":    set_type,
                    "Blhr(원본)": fmt_hhmm(blhr_h),
                    "Blhr_h":     blhr_h,
                    "교관시간_h":  instr_h,
                    "교관시간":   fmt_hhmm(instr_h),
                })

    return pd.DataFrame(detail_rows)

# ── 전체 승무원 Roster 파싱 — DH/OBCA/OBFO 감지 + OBCA/OBFO Blhr 합산 ───────
def parse_allcrew_roster(uploaded):
    """
    전체 승무원 Roster에서:
    - DH 탑승자 → (date_str, flight_no) 기준 연장/야간 제외 목록
    - OBCA/OBFO 탑승자 → 동일하게 제외 목록 + 승무원별 Blhr 합산 (총비행시간 차감)
    반환: dh_ob_exclude dict, ob_sum DataFrame (Crew Code, OBCA_OBFO)
    """
    raw, name_indices = _get_crew_sections(uploaded)
    dh_ob_exclude = {}   # (date_str, flight_no) -> set of names
    ob_hours      = {}   # crew_name -> 누적 OBCA/OBFO Blhr(h)

    for idx_i, name_idx in enumerate(name_indices[:-1]):
        crew_name = str(raw.iloc[name_idx, 0]).replace(":", "").strip()
        data = _parse_crew_data(raw, name_idx, name_indices[idx_i + 1])
        if data is None:
            continue

        for gid, grp in data.groupby("group_id"):
            if grp["Pairing_ff"].isna().all(): continue

            dc_vals  = grp["DC"].astype(str).str.strip().str.upper()
            pos_vals = grp["Pos"].astype(str).str.strip().str.upper()
            is_dh = dc_vals.eq("DH").any()
            is_ob = pos_vals.isin(["OBFO","OBCA"]).any()

            if not (is_dh or is_ob):
                continue

            flights = grp[grp["Activity"].apply(is_actual_flight)]
            for _, frow in flights.iterrows():
                date_str = str(frow["Date_ff"]).strip().upper()
                activity = str(frow["Activity"]).strip()
                m = re.search(r'\d+', activity)
                if not m: continue
                flight_no = str(int(m.group())).lstrip("0") or "0"
                key = (date_str, flight_no)
                dh_ob_exclude.setdefault(key, set()).add(crew_name)

                # OBCA/OBFO만 Blhr 합산 (총비행시간 차감용)
                if is_ob:
                    blhr_h = parse_hhmm(frow["Blhr"])
                    ob_hours[crew_name] = ob_hours.get(crew_name, 0.0) + blhr_h

    ob_sum = pd.DataFrame([
        {"Crew Code": name, "OBCA_OBFO": hrs}
        for name, hrs in ob_hours.items()
    ]) if ob_hours else pd.DataFrame(columns=["Crew Code","OBCA_OBFO"])

    return dh_ob_exclude, ob_sum

# ═══════════════════════════════════════════
# FltReport 수당 계산
# ═══════════════════════════════════════════
def combine_dt(date_str, t):
    if pd.isna(t) or t is None: return None
    if isinstance(date_str, (datetime.datetime, pd.Timestamp)):
        base = date_str
    else:
        s = str(date_str).strip()
        for fmt in ("%d%b%y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
            try:
                base = datetime.datetime.strptime(s, fmt)
                break
            except ValueError:
                continue
        else:
            return None
    return datetime.datetime(base.year, base.month, base.day, t.hour, t.minute, t.second)

def calc_night(ci, co):
    if not ci or not co: return 0.0
    total=0.0; d=ci.date()
    for delta in [-1,0,1]:
        ns=datetime.datetime.combine(d+timedelta(days=delta), time(22,0))
        ne=datetime.datetime.combine(d+timedelta(days=delta+1), time(6,0))
        o1=max(ci,ns); o2=min(co,ne)
        if o2>o1: total+=(o2-o1).total_seconds()/3600
    return total

def elapsed_hours(start, end):
    if not start or not end: return 0.0
    return max(0.0, (end-start).total_seconds()/3600)

def blhrs_decimal(t):
    if pd.isna(t) or t is None: return 0.0
    if isinstance(t, time): return t.hour + t.minute/60 + t.second/3600
    return 0.0

def is_oal_flight(flight):
    return "OAL" in str(flight).strip().upper()

def normalize_flight(flight):
    if is_oal_flight(flight):
        return None
    try:
        return str(int(float(str(flight)))).zfill(4)
    except:
        return str(flight).strip()

def adjust_ci_co_dates(atd_utc, ci_utc, co_utc):
    if ci_utc and atd_utc and ci_utc > atd_utc:
        ci_utc -= timedelta(days=1)
    if co_utc and ci_utc and co_utc < ci_utc:
        co_utc += timedelta(days=1)
    return ci_utc, co_utc

def process_flt(df, target_month, target_year, dh_ob_exclude=None):
    sum_rows, det_rows = [], []
    grouped = {}

    for _, row in df.iterrows():
        crew_str = row["운항 Crew"]
        if pd.isna(crew_str): continue

        raw_date = row["Date"]
        if isinstance(raw_date, (datetime.datetime, pd.Timestamp)):
            date_str = raw_date.strftime("%d%b%y").upper()
        else:
            date_str = str(raw_date).strip()

        flight = normalize_flight(row["Flight"])
        if not flight:
            continue

        atd_utc = combine_dt(date_str, row["ATD"])
        if not atd_utc: continue
        ata_utc = combine_dt(date_str, row["ATA"])
        ci_utc  = combine_dt(date_str, row["운항 C/I"])
        co_utc  = combine_dt(date_str, row["운항 C/O"])
        ci_utc, co_utc = adjust_ci_co_dates(atd_utc, ci_utc, co_utc)

        atd_kst = atd_utc + KST_OFFSET
        ata_kst = (ata_utc + KST_OFFSET) if ata_utc else None
        ci_kst  = (ci_utc  + KST_OFFSET) if ci_utc  else None
        co_kst  = (co_utc  + KST_OFFSET) if co_utc  else None

        if ata_kst and ata_kst < atd_kst: ata_kst += timedelta(days=1)
        if co_kst and ci_kst and co_kst < ci_kst: co_kst += timedelta(days=1)

        night_ref_kst = ci_kst if ci_kst else atd_kst

        atd_in_month   = (atd_kst.month == target_month and atd_kst.year == target_year)
        night_in_month = (night_ref_kst.month == target_month and night_ref_kst.year == target_year)

        if not atd_in_month and not night_in_month:
            continue

        bl    = blhrs_decimal(row["Bl Hrs"])
        night = calc_night(ci_kst, co_kst) if night_in_month else 0.0
        route = f"{row['From']}→{row['To']}"

        key = (date_str, flight, str(crew_str).strip())
        if key not in grouped:
            grouped[key] = {
                "date_str": date_str,
                "flight": flight,
                "crew_str": str(crew_str).strip(),
                "routes": [],
                "atd_kst": atd_kst,
                "ata_kst": ata_kst,
                "ci_kst": ci_kst,
                "co_kst": co_kst,
                "atd_in_month": atd_in_month,
                "seg_hours": 0.0,
                "night": 0.0,
                "p3_bl": 0.0,
            }
        g = grouped[key]
        g["routes"].append(route)
        if atd_kst and (not g["atd_kst"] or atd_kst < g["atd_kst"]):
            g["atd_kst"] = atd_kst
        if ata_kst and (not g["ata_kst"] or ata_kst > g["ata_kst"]):
            g["ata_kst"] = ata_kst
        if ci_kst and (not g["ci_kst"] or ci_kst < g["ci_kst"]):
            g["ci_kst"] = ci_kst
        if co_kst and (not g["co_kst"] or co_kst > g["co_kst"]):
            g["co_kst"] = co_kst
        if atd_in_month:
            g["seg_hours"] += elapsed_hours(atd_kst, ata_kst)
            if flight in ["0151", "0152"]:
                g["p3_bl"] += bl
        g["night"] += night

    for g in grouped.values():
        flight       = g["flight"]
        date_str     = g["date_str"]
        atd_in_month = g["atd_in_month"]
        night        = g["night"]
        ot           = max(0.0, g["seg_hours"] - 8) if atd_in_month else 0.0
        p3           = g["p3_bl"] if atd_in_month else 0.0

        if night == 0 and ot == 0 and p3 == 0:
            continue

        route    = " + ".join(g["routes"])
        dh_key   = (date_str, flight.lstrip("0") or "0")
        excluded = (dh_ob_exclude or {}).get(dh_key, set())

        for name in g["crew_str"].split():
            name = name.strip()
            if not name: continue
            if name in excluded: continue

            sum_rows.append({"이름": name, "night": night, "ot": ot, "p3": p3})
            det_rows.append({
                "이름":     name,
                "날짜":     g["atd_kst"].strftime("%m/%d"),
                "편명":     flight,
                "구간":     route,
                "ATD(KST)": fmt_time(g["atd_kst"]),
                "ATA(KST)": fmt_time(g["ata_kst"]),
                "CI(KST)":  fmt_time(g["ci_kst"]),
                "CO(KST)":  fmt_time(g["co_kst"]),
                "야간(h)":  night,
                "연장(h)":  ot,
                "3P(h)":    p3,
            })

    if not sum_rows: return pd.DataFrame(), pd.DataFrame()
    s = pd.DataFrame(sum_rows)
    summary = (s.groupby("이름")
                .agg(야간=("night","sum"), 연장=("ot","sum"), P3=("p3","sum"))
                .reset_index().sort_values("이름").reset_index(drop=True))
    detail = (pd.DataFrame(det_rows)
              .sort_values(["이름","날짜","편명"])
              .reset_index(drop=True))
    return summary, detail

# ═══════════════════════════════════════════
# 엑셀 빌드
# ═══════════════════════════════════════════
def build_excel(flt_sum, flt_det, calc_df, instr_det, target_year, target_month):
    wb = Workbook()
    bd=make_border(); center=Alignment(horizontal="center",vertical="center")
    left=Alignment(horizontal="left",vertical="center")
    cf=Font(name="Arial",size=10); tf=Font(name="Arial",bold=True,size=10)
    tbg=PatternFill("solid",fgColor="FFF2CC")
    sbf=Font(name="Arial",bold=True,color="FFFFFF",size=10)
    sbbg=PatternFill("solid",fgColor="2E75B6")
    redf=Font(name="Arial",bold=True,size=10,color="C00000")
    redbg=PatternFill("solid",fgColor="FCEAEA")
    label=f"{target_year}년 {target_month}월"

    def title_row(ws, text, ncols, note=None):
        col_letter = get_column_letter(ncols)
        ws.merge_cells(f"A1:{col_letter}1")
        c=ws["A1"]; c.value=text
        c.font=Font(name="Arial",bold=True,size=13); c.alignment=center
        ws.row_dimensions[1].height=28
        if note:
            ws.merge_cells(f"A2:{col_letter}2")
            c2=ws["A2"]; c2.value=note
            c2.font=Font(name="Arial",italic=True,size=9,color="777777")
            c2.alignment=left

    def tot_row(ws, row_n, ncols, vals_dict):
        for col in range(1,ncols+1):
            c=ws.cell(row=row_n,column=col); c.font=tf; c.fill=tbg; c.alignment=center; c.border=bd
        for col, val in vals_dict.items():
            ws.cell(row=row_n,column=col).value=val

    # ── 시트1: 수당 요약 ─────────────────────
    ws1=wb.active; ws1.title="수당 요약"
    title_row(ws1, f"{label} 운항 수당 정산표 — 요약", 5,
              "※ 야간: 22:00~06:00(KST) C/I기준 귀속 | 연장: 일 8시간 초과(ATD기준) | 3P: 편명 0151·0152 | OAL·DH·OBCA/OBFO 제외")
    style_hdr(ws1, 3, ["No","이름","야간 시간","연장 시간","3P 시간"])
    for i,row in flt_sum.iterrows():
        r=i+4
        for col,val in enumerate([i+1,row["이름"],fmt_hhmm(row["야간"]),fmt_hhmm(row["연장"]),fmt_hhmm(row["P3"])],1):
            c=ws1.cell(row=r,column=col,value=val); c.font=cf; c.border=bd
            c.alignment=left if col==2 else center
    tr=len(flt_sum)+4
    ws1.merge_cells(f"A{tr}:B{tr}")
    tot_row(ws1,tr,5,{1:"합계",3:fmt_hhmm(flt_sum["야간"].sum()),4:fmt_hhmm(flt_sum["연장"].sum()),5:fmt_hhmm(flt_sum["P3"].sum())})
    for i,w in enumerate([5,14,11,11,11],1): ws1.column_dimensions[get_column_letter(i)].width=w
    ws1.freeze_panes="A4"

    # ── 시트2: 개인별 상세 ───────────────────
    ws2=wb.create_sheet("개인별 상세")
    title_row(ws2, f"{label} 운항 수당 정산표 — 개인별 비행 상세", 11)
    style_hdr(ws2,2,["이름","날짜","편명","구간","ATD(KST)","ATA(KST)","CI(KST)","CO(KST)","야간(h)","연장(h)","3P(h)"])
    names=sorted(flt_det["이름"].unique()); cur=3; ci=0; fills=["FFFFFF","F7FBFF"]
    for name in names:
        grp=flt_det[flt_det["이름"]==name].reset_index(drop=True)
        rf=PatternFill("solid",fgColor=fills[ci%2]); ci+=1
        for _,dr in grp.iterrows():
            for col,key in enumerate(["이름","날짜","편명","구간","ATD(KST)","ATA(KST)","CI(KST)","CO(KST)","야간(h)","연장(h)","3P(h)"],1):
                val=dr[key]
                if key in ["야간(h)","연장(h)","3P(h)"]: val=fmt_hhmm(val)
                c=ws2.cell(row=cur,column=col,value=val); c.font=cf; c.fill=rf; c.border=bd
                c.alignment=left if col in [1,4] else center
            cur+=1
        sv=[name+" 소계","","","","","","","",fmt_hhmm(grp["야간(h)"].sum()),fmt_hhmm(grp["연장(h)"].sum()),fmt_hhmm(grp["3P(h)"].sum())]
        for col,val in enumerate(sv,1):
            c=ws2.cell(row=cur,column=col,value=val); c.font=sbf; c.fill=sbbg; c.alignment=center; c.border=bd
        cur+=1
    for col in range(1,12):
        c=ws2.cell(row=cur,column=col); c.font=tf; c.fill=tbg; c.alignment=center; c.border=bd
    ws2.cell(row=cur,column=1).value="전체 합계"
    ws2.cell(row=cur,column=9).value=fmt_hhmm(flt_det["야간(h)"].sum())
    ws2.cell(row=cur,column=10).value=fmt_hhmm(flt_det["연장(h)"].sum())
    ws2.cell(row=cur,column=11).value=fmt_hhmm(flt_det["3P(h)"].sum())
    for i,w in enumerate([14,8,8,12,10,10,10,10,10,10,10],1): ws2.column_dimensions[get_column_letter(i)].width=w
    ws2.freeze_panes="A3"

    # ── 시트3: 총비행시간 ────────────────────
    ws3=wb.create_sheet("총비행시간")
    hdrs3=["No","이름","Block(h)","DHC(h)","DHC×50%(h)","OBCA/OBFO(h)","총비행시간(h)"]
    title_row(ws3,f"{label} 개인별 총 비행시간",len(hdrs3),"※ 총비행시간 = Block + DHC×50% - OBCA/OBFO")
    style_hdr(ws3,3,hdrs3)
    ds=calc_df.sort_values("Crew Code").reset_index(drop=True)
    for i,row in ds.iterrows():
        r=i+4
        for col,val in enumerate([i+1,row["Crew Code"],fmt_hhmm(row["Block"]),fmt_hhmm(row["DHC"]),
                                   fmt_hhmm(row["DHC_50"]),fmt_hhmm(row["OBCA_OBFO"]),fmt_hhmm(row["Total_Flt"])],1):
            c=ws3.cell(row=r,column=col,value=val); c.font=cf; c.border=bd
            c.alignment=left if col==2 else center
    tr3=len(ds)+4; ws3.merge_cells(f"A{tr3}:B{tr3}")
    tot_row(ws3,tr3,len(hdrs3),{1:"합계",3:fmt_hhmm(ds["Block"].sum()),4:fmt_hhmm(ds["DHC"].sum()),
             5:fmt_hhmm(ds["DHC_50"].sum()),6:fmt_hhmm(ds["OBCA_OBFO"].sum()),7:fmt_hhmm(ds["Total_Flt"].sum())})
    for i,w in enumerate([5,14,11,11,11,13,13],1): ws3.column_dimensions[get_column_letter(i)].width=w
    ws3.freeze_panes="A4"

    # ── 시트4: DAYOFF 미달 ───────────────────
    ws4=wb.create_sheet("DAYOFF 8회 미만")
    hdrs4=["No","이름","DAYOFF 횟수","비고"]
    title_row(ws4,f"{label} DAYOFF 8회 미만자",len(hdrs4))
    style_hdr(ws4,2,hdrs4)
    u8=calc_df[calc_df["Dayoff_Under8"]].sort_values("Crew Code").reset_index(drop=True)
    if len(u8)==0:
        ws4.merge_cells("A3:D3")
        c=ws4["A3"]; c.value="8회 미만자 없음"
        c.font=Font(name="Arial",italic=True,size=10,color="777777"); c.alignment=center
    else:
        for i,row in u8.iterrows():
            r=i+3
            for col,val in enumerate([i+1,row["Crew Code"],int(row["Dayoff"]),f"기준 미달 ({int(row['Dayoff'])}회 / 8회)"],1):
                c=ws4.cell(row=r,column=col,value=val); c.font=redf; c.fill=redbg; c.border=bd
                c.alignment=left if col in [2,4] else center
    for i,w in enumerate([5,14,14,24],1): ws4.column_dimensions[get_column_letter(i)].width=w
    ws4.freeze_panes="A3"

    # ── 시트5: 교관 수당 요약 ────────────────
    ws5=wb.create_sheet("교관수당 요약")
    hdrs5=["No","이름","1set 합계","2set 합계","3P 합계","교관시간 총계"]
    title_row(ws5,f"{label} 교관 비행 수당 요약",len(hdrs5),
              "※ 1set=BH전체(DAD·BKK·HKG·NRT) | 2set=BH×1/2(DAC·LAX·EWR·SFO·IAD) | 3P=BH×1/3(HNL)")
    style_hdr(ws5,3,hdrs5)
    if not instr_det.empty:
        instr_sum=(instr_det.groupby("이름").agg(
            s1=("교관시간_h", lambda x: x[instr_det.loc[x.index,"Set구분"]=="1set"].sum()),
            s2=("교관시간_h", lambda x: x[instr_det.loc[x.index,"Set구분"]=="2set"].sum()),
            s3=("교관시간_h", lambda x: x[instr_det.loc[x.index,"Set구분"]=="3P"].sum()),
            tot=("교관시간_h","sum")
        ).reset_index().sort_values("이름").reset_index(drop=True))
        for i,row in instr_sum.iterrows():
            r=i+4
            for col,val in enumerate([i+1,row["이름"],fmt_hhmm(row["s1"]),fmt_hhmm(row["s2"]),fmt_hhmm(row["s3"]),fmt_hhmm(row["tot"])],1):
                c=ws5.cell(row=r,column=col,value=val); c.font=cf; c.border=bd
                c.alignment=left if col==2 else center
        tr5=len(instr_sum)+4; ws5.merge_cells(f"A{tr5}:B{tr5}")
        tot_row(ws5,tr5,len(hdrs5),{1:"합계",
            3:fmt_hhmm(instr_det[instr_det["Set구분"]=="1set"]["교관시간_h"].sum()),
            4:fmt_hhmm(instr_det[instr_det["Set구분"]=="2set"]["교관시간_h"].sum()),
            5:fmt_hhmm(instr_det[instr_det["Set구분"]=="3P"]["교관시간_h"].sum()),
            6:fmt_hhmm(instr_det["교관시간_h"].sum())})
    for i,w in enumerate([5,14,12,12,12,14],1): ws5.column_dimensions[get_column_letter(i)].width=w
    ws5.freeze_panes="A4"

    # ── 시트6: 교관 수당 상세 ────────────────
    ws6=wb.create_sheet("교관수당 상세")
    hdrs6=["이름","날짜(KST)","DC","편명","From","To","Set구분","Blhr(원본)","교관시간","산출방식"]
    title_row(ws6,f"{label} 교관 비행 수당 상세",len(hdrs6))
    style_hdr(ws6,2,hdrs6)
    cur6=3; ci6=0; fills6=["FFFFFF","EDF4FF"]
    if not instr_det.empty:
        names6=sorted(instr_det["이름"].unique())
        for name in names6:
            grp=instr_det[instr_det["이름"]==name].reset_index(drop=True)
            rf=PatternFill("solid",fgColor=fills6[ci6%2]); ci6+=1
            for _,dr in grp.iterrows():
                st=dr["Set구분"]
                if st=="1set":   formula="BH × 1 (1set)"
                elif st=="2set": formula="BH × 1/2 (2set)"
                else:            formula="BH × 1/3 (3P)"
                vals=[dr["이름"],dr["날짜(KST)"],dr["DC"],dr["편명"],dr["From"],dr["To"],
                      st,dr["Blhr(원본)"],dr["교관시간"],formula]
                for col,val in enumerate(vals,1):
                    c=ws6.cell(row=cur6,column=col,value=val); c.font=cf; c.fill=rf; c.border=bd
                    c.alignment=left if col in [1,4,10] else center
                cur6+=1
            sv=[name+" 소계","","","","","","",
                "",fmt_hhmm(grp["교관시간_h"].sum()),""]
            for col,val in enumerate(sv,1):
                c=ws6.cell(row=cur6,column=col,value=val); c.font=sbf; c.fill=sbbg; c.alignment=center; c.border=bd
            cur6+=1
        for col in range(1,len(hdrs6)+1):
            c=ws6.cell(row=cur6,column=col); c.font=tf; c.fill=tbg; c.alignment=center; c.border=bd
        ws6.cell(row=cur6,column=1).value="전체 합계"
        ws6.cell(row=cur6,column=9).value=fmt_hhmm(instr_det["교관시간_h"].sum())
    for i,w in enumerate([14,9,8,10,6,6,8,10,10,16],1): ws6.column_dimensions[get_column_letter(i)].width=w
    ws6.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf

# ═══════════════════════════════════════════
# Streamlit UI
# ═══════════════════════════════════════════
st.set_page_config(page_title="운항 수당 정산기", page_icon="✈️", layout="centered")
st.title("✈️ 운항 수당 정산기")
st.caption("5개 파일 업로드 → 월 선택 → 정산 실행 → 엑셀 다운로드 (시트 6개)")
st.divider()

c1, c2 = st.columns(2)
with c1:
    st.markdown("**📂 FltReport.xlsx**")
    st.caption("매월 2일 오전 9:30 메일 수신 · 발송처: yp-report@airpremia.com · 제목: PDC_FLT_Report")
    flt_file = st.file_uploader("FltReport.xlsx", type=["xlsx"], label_visibility="collapsed", key="up_flt")

    st.markdown("**📂 월DHC_DAYOFF_총비행시간.xlsx**")
    st.caption("PDC → Reports → Counter report → Period 설정 → All in FD → Counter에 Block · Day off · DHC 선택 → 추출")
    dhc_file = st.file_uploader("월DHC_DAYOFF_총비행시간.xlsx", type=["xlsx"], label_visibility="collapsed", key="up_dhc")

    st.markdown("**📂 OBCA.xlsx (OBCA/OBFO 전용)**")
    st.caption("PDC → Reports → Experience → 기간 선택 및 출력")
    ob_file = st.file_uploader("OBCA.xlsx", type=["xlsx"], label_visibility="collapsed", key="up_ob")

with c2:
    st.markdown("**📂 Roster.xlsx (교관수당용)**")
    st.caption("PDC → Crew roster → Position(LIP·LCP·DLCP) → Period 설정(⚠️ Time mode=UTC 기준이므로 전월 말일~익월 1일 하루씩 여유 있게, 예: 30Jun26~01Aug26) → Section(Schedule) → Time mode(UTC) → 추출")
    rost_file = st.file_uploader("Roster.xlsx (교관수당용)", type=["xlsx"], label_visibility="collapsed", key="up_rost")

    st.markdown("**📂 Roster.xlsx (전체 승무원)**")
    st.caption("PDC → Crew roster → Position(All in FD) → Period 설정(⚠️ Time mode=UTC 기준이므로 전월 말일~익월 1일 하루씩 여유 있게) → Counter에 Pairing · duty code · working position · check in(local) · check out (local) · Activity · From · To · A/C HOTEL · Block · FDP TIME 선택 → 추출")
    allcrew_file = st.file_uploader("Roster.xlsx (전체 승무원)", type=["xlsx"], label_visibility="collapsed", key="up_allcrew")

all_uploaded = flt_file and dhc_file and ob_file and rost_file and allcrew_file

if all_uploaded:
    try:
        flt_df = pd.read_excel(flt_file)
        required = {"Date","Flight","ATD","ATA","Bl Hrs","운항 C/I","운항 C/O","운항 Crew"}
        missing = required - set(flt_df.columns)
        if missing: st.error(f"FltReport 필수 열 없음: {missing}"); st.stop()

        # ── 램프리턴/불이착 자동 보정 ──────────────────────────
        flt_df, ramp_log = normalize_ramp_returns(flt_df)
        if ramp_log:
            with st.expander(f"🛬 램프리턴/회항 자동 보정 내역 ({len(ramp_log)}건)"):
                for date_label, flight, origin, desc in ramp_log:
                    st.write(f"**{date_label} 편명 {flight} ({origin})**: {desc}")
                st.caption("※ 목적지 자체가 바뀌는 진짜 다이버트(대체공항 착륙 후 별도 이동편)는 "
                           "원본 데이터에 이동편 기록이 없는 경우가 있어 자동 보정 대상에서 제외됩니다. "
                           "이런 경우는 별도로 확인해주세요.")

        flt_df["date_parsed"] = pd.to_datetime(flt_df["Date"], format="%d%b%y")
        available = sorted(flt_df[flt_df["운항 Crew"].notna()]["date_parsed"].dt.to_period("M").unique(), reverse=True)

        base_df = parse_dhc_file(dhc_file)

        # 전체 승무원 Roster → DH/OBCA/OBFO 제외 목록 + OBCA/OBFO Blhr 합산(자동 감지분)
        dh_ob_exclude, ob_sum_roster = parse_allcrew_roster(allcrew_file)

        # 별도 OBCA.xlsx → OBCA/OBFO Block Hours 합산(전용 파일분)
        ob_sum_file = parse_obca_file(ob_file)

        # 두 출처 합산 (승무원별로 겹치지 않는다는 전제하에 단순 합산)
        ob_sum = merge_ob_sums(ob_sum_roster, ob_sum_file)

        calc_df = calc_summary(base_df, ob_sum)

        # 편명 → Set구분 매핑 (GMP 등 국내 재배치 구간 보정용) + 교관수당 Roster 파싱
        flight_set_map = build_flight_set_map(rost_file)
        instr_det_all = parse_roster_file(rost_file, flight_set_map=flight_set_map)

        n_instr    = instr_det_all["이름"].nunique() if not instr_det_all.empty else 0
        n_dh_pairs = len(dh_ob_exclude)
        dh_names   = set(n for names in dh_ob_exclude.values() for n in names)
        n_ob       = len(ob_sum)
        st.success(f"✅ 5개 파일 로드 완료 — FltReport {len(flt_df):,}행 · 승무원 {len(base_df)}명 · 교관 {n_instr}명 · DH/OBCA/OBFO 제외 {len(dh_names)}명({n_dh_pairs}건) · OBCA/OBFO 차감 {n_ob}명")

        selected_str = st.selectbox("📅 정산할 월 선택", [str(p) for p in available])
        sel = pd.Period(selected_str, freq="M"); target_year, target_month = sel.year, sel.month

        with st.expander("ℹ️ 계산 기준 보기"):
            st.markdown("""
| 항목 | 기준 |
|------|------|
| **램프리턴/회항 보정** | From==To(같은 편명 회항) 행 자동 감지 → 회항 소요시간과 무관하게 재출발편과 병합 (ATD는 회항편 최초 출발시각 기준 → 연장시간에 회항 경과시간 포함) |
| **야간** | 운항 C/I~C/O(KST) 중 22:00~06:00 겹치는 시간 / C/I KST 기준 월 귀속 |
| **연장** | ATD~ATA(KST) 8시간 초과분 / ATD KST 기준 월 귀속 / 같은 편명 분할행은 구간 시간 합산 |
| **3P** | 편명 0151·0152 Bl Hrs / ATD KST 기준 월 귀속 |
| **OAL** | 편명에 OAL 포함된 행 제외 |
| **DH 제외** | 전체 Roster DC='DH' 자동 감지 → 연장·야간 제외 |
| **OBCA/OBFO 제외** | 전체 Roster Pos='OBCA'/'OBFO' 자동 감지 → 연장·야간 제외 |
| **OBCA/OBFO 차감** | 전체 Roster 자동감지분 + 별도 OBCA.xlsx 합산분 → 총비행시간 차감 |
| **총비행시간** | Block + DHC×50% - OBCA/OBFO(Blhr 합산) |
| **DAYOFF 미달** | 월 DAYOFF 8회 미만 |
| **교관수당 1set** | BH 전체 (DAD·BKK·HKG·NRT) |
| **교관수당 2set** | BH × 1/2 (DAC·LAX·EWR·SFO·IAD) |
| **교관수당 3P** | BH × 1/3 (HNL) |
| **교관 DC** | LIP·LCP·DLCP·I·I* |
            """)

        if dh_ob_exclude:
            with st.expander(f"✈️ DH/OBCA/OBFO 자동 감지 현황 ({len(dh_ob_exclude)}건)"):
                for (date_str, flt_no), names in sorted(dh_ob_exclude.items()):
                    st.write(f"**{date_str} 편명 {flt_no}**: {', '.join(sorted(names))}")

        if st.button("🚀 정산 실행", type="primary", use_container_width=True):
            with st.spinner("계산 중..."):
                flt_sum, flt_det = process_flt(flt_df, target_month, target_year, dh_ob_exclude)
                instr_det = parse_roster_file(rost_file, target_month, target_year, flight_set_map=flight_set_map)

            if flt_sum.empty:
                st.warning("FltReport에서 해당 월 데이터가 없습니다.")
            else:
                st.success(f"✅ {target_year}년 {target_month}월 정산 완료")

                tab1,tab2,tab3,tab4,tab5,tab6 = st.tabs(
                    ["📊 수당 요약","📋 개인별 상세","⏱ 총비행시간","📅 DAYOFF 미달","🎓 교관수당 요약","🎓 교관수당 상세"])

                with tab1:
                    d=flt_sum.copy()
                    for k in ["야간","연장","P3"]: d[k]=d[k].apply(fmt_hhmm)
                    d.index+=1; d.columns=["이름","야간 시간","연장 시간","3P 시간"]
                    st.dataframe(d,use_container_width=True,height=360)
                    c1,c2,c3=st.columns(3)
                    c1.metric("야간",fmt_hhmm(flt_sum["야간"].sum()))
                    c2.metric("연장",fmt_hhmm(flt_sum["연장"].sum()))
                    c3.metric("3P",fmt_hhmm(flt_sum["P3"].sum()))

                with tab2:
                    nl=["전체"]+sorted(flt_det["이름"].unique().tolist())
                    sn=st.selectbox("승무원 선택",nl)
                    vw=flt_det if sn=="전체" else flt_det[flt_det["이름"]==sn]
                    d2=vw.copy()
                    for col in ["야간(h)","연장(h)","3P(h)"]: d2[col]=d2[col].apply(fmt_hhmm)
                    st.dataframe(d2.reset_index(drop=True),use_container_width=True,height=380)

                with tab3:
                    d3=calc_df.sort_values("Crew Code").reset_index(drop=True).copy()
                    for col in ["Block","DHC","DHC_50","OBCA_OBFO","Total_Flt"]: d3[col]=d3[col].apply(fmt_hhmm)
                    d3=d3[["Crew Code","Block","DHC","DHC_50","OBCA_OBFO","Total_Flt","Dayoff"]]
                    d3.columns=["이름","Block(h)","DHC(h)","DHC×50%(h)","OBCA/OBFO(h)","총비행시간(h)","DAYOFF"]
                    d3.index+=1
                    st.dataframe(d3,use_container_width=True,height=380)

                with tab4:
                    u8=calc_df[calc_df["Dayoff_Under8"]].sort_values("Crew Code").reset_index(drop=True)
                    if len(u8)==0: st.success("8회 미만자 없음 ✅")
                    else:
                        st.error(f"⚠️ DAYOFF 8회 미만자 {len(u8)}명")
                        d4=u8[["Crew Code","Dayoff"]].copy(); d4.columns=["이름","DAYOFF 횟수"]; d4.index+=1
                        st.dataframe(d4,use_container_width=True)

                with tab5:
                    if instr_det.empty:
                        st.info("교관 비행 데이터 없음")
                    else:
                        instr_sum=(instr_det.groupby("이름").agg(
                            s1=("교관시간_h",lambda x:x[instr_det.loc[x.index,"Set구분"]=="1set"].sum()),
                            s2=("교관시간_h",lambda x:x[instr_det.loc[x.index,"Set구분"]=="2set"].sum()),
                            s3=("교관시간_h",lambda x:x[instr_det.loc[x.index,"Set구분"]=="3P"].sum()),
                            tot=("교관시간_h","sum")
                        ).reset_index().sort_values("이름").reset_index(drop=True))
                        d5=instr_sum.copy()
                        for col in ["s1","s2","s3","tot"]: d5[col]=d5[col].apply(fmt_hhmm)
                        d5.index+=1; d5.columns=["이름","1set 합계","2set 합계","3P 합계","교관시간 총계"]
                        st.dataframe(d5,use_container_width=True,height=360)
                        c1,c2,c3,c4=st.columns(4)
                        c1.metric("1set",fmt_hhmm(instr_det[instr_det["Set구분"]=="1set"]["교관시간_h"].sum()))
                        c2.metric("2set",fmt_hhmm(instr_det[instr_det["Set구분"]=="2set"]["교관시간_h"].sum()))
                        c3.metric("3P",fmt_hhmm(instr_det[instr_det["Set구분"]=="3P"]["교관시간_h"].sum()))
                        c4.metric("전체",fmt_hhmm(instr_det["교관시간_h"].sum()))

                with tab6:
                    if instr_det.empty:
                        st.info("교관 비행 데이터 없음")
                    else:
                        nl2=["전체"]+sorted(instr_det["이름"].unique().tolist())
                        sn2=st.selectbox("교관 선택",nl2,key="instr_sel")
                        vw2=instr_det if sn2=="전체" else instr_det[instr_det["이름"]==sn2]
                        d6=vw2[["이름","날짜(KST)","DC","편명","From","To","Set구분","Blhr(원본)","교관시간"]].copy()
                        d6.index+=1
                        st.dataframe(d6,use_container_width=True,height=400)
                        if sn2!="전체":
                            cc1,cc2,cc3=st.columns(3)
                            cc1.metric("1set",fmt_hhmm(vw2[vw2["Set구분"]=="1set"]["교관시간_h"].sum()))
                            cc2.metric("2set",fmt_hhmm(vw2[vw2["Set구분"]=="2set"]["교관시간_h"].sum()))
                            cc3.metric("3P",fmt_hhmm(vw2[vw2["Set구분"]=="3P"]["교관시간_h"].sum()))

                excel_buf = build_excel(flt_sum, flt_det, calc_df, instr_det, target_year, target_month)
                fname = f"{target_year}년{target_month:02d}월_운항정산.xlsx"
                st.download_button(
                    label="⬇️ 엑셀 다운로드 (6개 시트: 수당요약·개인상세·총비행시간·DAYOFF·교관요약·교관상세)",
                    data=excel_buf, file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)

    except Exception as e:
        st.error(f"오류 발생: {e}"); st.exception(e)
else:
    missing=[]
    if not flt_file:      missing.append("FltReport.xlsx")
    if not dhc_file:      missing.append("월DHC_DAYOFF_총비행시간.xlsx")
    if not ob_file:       missing.append("OBCA.xlsx")
    if not rost_file:     missing.append("Roster.xlsx (교관수당용)")
    if not allcrew_file:  missing.append("Roster.xlsx (전체 승무원)")
    st.info(f"아래 파일을 모두 업로드해주세요: {' · '.join(missing)}")
